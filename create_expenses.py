import sys
import io
if isinstance(sys.stdout, io.TextIOWrapper):
    sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()

# ── สีและ style ───────────────────────────────────────────────────────────────
สีรายรับ  = "1a6e3f"   # เขียวเข้ม
สีรายจ่าย = "8b1a1a"   # แดงเข้ม
สีสรุป    = "1a3a6e"   # น้ำเงินเข้ม
สีหัวรายรับ  = PatternFill("solid", fgColor="d4edda")
สีหัวรายจ่าย = PatternFill("solid", fgColor="f8d7da")
สีหัวสรุป    = PatternFill("solid", fgColor="cce5ff")
สีแถวลง      = PatternFill("solid", fgColor="f9f9f9")

เส้น = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)

def จัดหัวคอลัมน์(ws, headers, สีหัว, สีข้อความ):
    for col, ชื่อ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=ชื่อ)
        cell.font      = Font(bold=True, color=สีข้อความ, size=11)
        cell.fill      = สีหัว
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = เส้น

def จัดความกว้าง(ws, กว้าง):
    for col, w in enumerate(กว้าง, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

def ใส่แถว(ws, row_num, values, number_format="฿#,##0.00"):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border    = เส้น
        cell.alignment = Alignment(vertical='center')
        if col == 3:   # คอลัมน์จำนวนเงิน
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal='right', vertical='center')
        if row_num % 2 == 0:
            cell.fill = สีแถวลง

# ── Sheet 1: รายรับ ───────────────────────────────────────────────────────────
ws_รับ = wb.active
ws_รับ.title = "รายรับ"
ws_รับ.freeze_panes = "A2"
จัดหัวคอลัมน์(ws_รับ, ["วันที่", "รายการ", "จำนวนเงิน", "หมวดหมู่"], สีหัวรายรับ, สีรายรับ)
จัดความกว้าง(ws_รับ, [14, 32, 16, 18])

ข้อมูลรายรับ = [
    (date(2026, 5, 1),  "เงินเดือน",   45000, "เงินเดือน"),
    (date(2026, 5, 5),  "รายได้พิเศษ", 8000,  "รายได้อื่น"),
    (date(2026, 5, 10), "เงินคืนภาษี", 3200,  "รายได้อื่น"),
]
for i, แถว in enumerate(ข้อมูลรายรับ, 2):
    ใส่แถว(ws_รับ, i, แถว)
    ws_รับ.cell(i, 1).number_format = "YYYY-MM-DD"

# ── Sheet 2: รายจ่าย ──────────────────────────────────────────────────────────
ws_จ่าย = wb.create_sheet("รายจ่าย")
ws_จ่าย.freeze_panes = "A2"
จัดหัวคอลัมน์(ws_จ่าย, ["วันที่", "รายการ", "จำนวนเงิน", "หมวดหมู่"], สีหัวรายจ่าย, สีรายจ่าย)
จัดความกว้าง(ws_จ่าย, [14, 32, 16, 18])

ข้อมูลรายจ่าย = [
    (date(2026, 5, 1),  "ค่าเช่าบ้าน",            8500, "ที่อยู่อาศัย"),
    (date(2026, 5, 3),  "ค่าอาหาร",               4200, "อาหาร"),
    (date(2026, 5, 7),  "ค่าน้ำค่าไฟ",             1800, "สาธารณูปโภค"),
    (date(2026, 5, 8),  "ค่าเดินทาง",              1200, "การเดินทาง"),
    (date(2026, 5, 9),  "ค่าโทรศัพท์อินเทอร์เน็ต",  699, "การสื่อสาร"),
]
for i, แถว in enumerate(ข้อมูลรายจ่าย, 2):
    ใส่แถว(ws_จ่าย, i, แถว)
    ws_จ่าย.cell(i, 1).number_format = "YYYY-MM-DD"

# ── Sheet 3: สรุป (formula อัตโนมัติ — ห้ามแก้มือ) ──────────────────────────
ws_สรุป = wb.create_sheet("สรุป")
ws_สรุป.column_dimensions["A"].width = 22
ws_สรุป.column_dimensions["B"].width = 20

# หัว
ws_สรุป["A1"] = "สรุปรายรับรายจ่าย"
ws_สรุป["A1"].font = Font(bold=True, size=14, color=สีสรุป)
ws_สรุป["A1"].fill = สีหัวสรุป
ws_สรุป["B1"] = "⚠ อัปเดตอัตโนมัติ ห้ามแก้มือ"
ws_สรุป["B1"].font = Font(italic=True, color="888888", size=9)
ws_สรุป.row_dimensions[1].height = 24

# แถวสรุป + formula ที่ Excel คำนวณอัตโนมัติ
แถวสรุป = [
    (3,  "รายรับรวม",    "=SUM(รายรับ!C2:C10000)",          "d4edda", สีรายรับ),
    (4,  "รายจ่ายรวม",   "=SUM(รายจ่าย!C2:C10000)",         "f8d7da", สีรายจ่าย),
    (5,  "ยอดคงเหลือ",   "=B3-B4",                          "cce5ff", สีสรุป),
    (7,  "จำนวนรายรับ",  "=COUNTA(รายรับ!B2:B10000)",       "f0f0f0", "333333"),
    (8,  "จำนวนรายจ่าย", "=COUNTA(รายจ่าย!B2:B10000)",      "f0f0f0", "333333"),
]
for row, label, formula, bg, fg in แถวสรุป:
    ws_สรุป.cell(row, 1, label).font  = Font(bold=True, color=fg)
    ws_สรุป.cell(row, 1).fill        = PatternFill("solid", fgColor=bg)
    ws_สรุป.cell(row, 1).border      = เส้น
    ws_สรุป.cell(row, 1).alignment   = Alignment(vertical='center')
    c = ws_สรุป.cell(row, 2, formula)
    c.font          = Font(bold=True, color=fg)
    c.fill          = PatternFill("solid", fgColor=bg)
    c.border        = เส้น
    c.alignment     = Alignment(horizontal='right', vertical='center')
    if row <= 5:
        c.number_format = "฿#,##0.00"

ws_สรุป["A6"] = ""   # เว้นบรรทัด

wb.save("expenses.xlsx")
print("สร้าง expenses.xlsx เรียบร้อย — 3 sheets: รายรับ, รายจ่าย, สรุป")
